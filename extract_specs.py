#!/usr/bin/env python3
"""
extract_specs.py — build specs.csv for fetch_sellsheets.py from the WPH
Master Price List workbook.

Pass 1 (always): scans the numbered category sheets, finds product blocks and
their SKU rows, writes:
    specs_master.csv   every row found: sheet, block, size_type, dimensions,
                       gsm, weight, case_pack, color
    blocks.txt         distinct product-block names (for building mapping.csv)

Pass 2 (if mapping.csv exists): joins blocks to Shopify handles and writes
    specs.csv          handle,size,gsm,dimensions,case_pack  (fetcher format)

mapping.csv format — block_match is a case-insensitive substring of the
price-list block header:
    handle,block_match
    martex-cam-towel-collection,Martex Cam Towel

Usage:
    python3 extract_specs.py "2026 WPH ABC Master Price List.xlsx" [--map mapping.csv]
"""

import argparse
import csv
import re
import sys
from collections import OrderedDict

from openpyxl import load_workbook

DIMS_RE = re.compile(
    r"^\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)(?:\s*[xX]\s*(\d+(?:\.\d+)?))?\s*$")

# description prefixes that name the size/type shown in the sell sheet table
TYPE_PREFIX_RE = re.compile(r"^([A-Za-z /&'-]+?)\s*-\s")

SIZEISH_RE = re.compile(
    r"\b(TWIN|FULL|QUEEN|KING|STANDARD|JUMBO|EURO|BODY|CAP)\b", re.I)
DEPTH_RE = re.compile(r'^\d+(?:\.\d+)?"$')


def split_desc(desc):
    """Description -> (item_type, size).
    'Pillow Case - Standard - Ultra Touch'   -> ('Pillow Case', 'Standard')
    'Twin Blanket - Sovereign Cotton'        -> ('', 'Twin Blanket')
    'Bath Towel - Martex Single Cam ...'     -> ('', 'Bath Towel')
    """
    parts = [norm(p) for p in desc.split(" - ") if norm(p)]
    if not parts:
        return "", ""
    seg1 = parts[0]
    seg2 = parts[1] if len(parts) > 1 else ""
    if seg2 and len(seg2) <= 20 and SIZEISH_RE.search(seg2):
        return seg1, seg2
    return "", seg1



def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def fmt_dims(raw):
    m = DIMS_RE.match(str(raw or ""))
    if m:
        parts = [g for g in m.groups() if g]
        return " x ".join(f'{p}"' for p in parts)
    return norm(raw)


def find_header(ws):
    """Locate the header row and map column names -> index."""
    for row in ws.iter_rows(min_row=1, max_row=8):
        names = {}
        for cell in row:
            v = norm(cell.value).replace("\n", " ").upper()
            if v and v not in names:  # keep first — headers repeat in the
                names[v] = cell.column - 1  # "Old SKU Info" area to the right
        if "DESCRIPTION" in names and any(k == "SIZE" or k.startswith("CUT SIZE")
                                          or k.startswith("FINISH SIZE")
                                          for k in names):
            def first(pred):
                for k, v in names.items():
                    if pred(k):
                        return v
                return None
            cols = {
                "size": first(lambda k: k.startswith("FINISH SIZE")) 
                        if first(lambda k: k.startswith("FINISH SIZE")) is not None
                        else names.get("SIZE"),
                "cut": first(lambda k: k.startswith("CUT SIZE")),
                "hem": first(lambda k: k.startswith("HEM")),
                "weight": first(lambda k: k == "LBS/DZ" or k.startswith("WEIGHT")),
                "weight_unit": ("lbs/dz" if "LBS/DZ" in names
                                else "oz" if any(k.startswith("WEIGHT (OZ") for k in names)
                                else ""),
                "gsm": names.get("GSM"),
                "desc": names.get("DESCRIPTION"),
                "color": first(lambda k: k.startswith("COLOR")),
                "small": names.get("SMALL CASE"),
                "large": names.get("LARGE CASE"),
                "case": names.get("CASE"),
            }
            return row[0].row, cols
    return None, None


def is_block_header(rowvals, cols):
    """A block header has long text in col A and no description data."""
    a = norm(rowvals[0] if rowvals else "")
    if len(a) < 30:
        return False
    desc = rowvals[cols["desc"]] if cols["desc"] is not None and cols["desc"] < len(rowvals) else None
    return desc in (None, "")


def cellv(rowvals, idx):
    if idx is None or idx >= len(rowvals):
        return ""
    return norm(rowvals[idx])


def size_type_from_desc(desc):
    m = TYPE_PREFIX_RE.match(desc)
    return norm(m.group(1)) if m else ""


def num(s):
    s = norm(s)
    if not s:
        return ""
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else str(f)
    except ValueError:
        return s


def extract(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    records = []
    for ws in wb.worksheets:
        if not ws.title.strip().isdigit():
            continue
        hdr_row, cols = find_header(ws)
        if not cols:
            continue
        block = ""
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            rowvals = list(row)
            if is_block_header(rowvals, cols):
                block = norm(rowvals[0])
                continue
            desc = cellv(rowvals, cols["desc"])
            gsm = num(cellv(rowvals, cols["gsm"]))
            if not desc:
                continue
            case = (num(cellv(rowvals, cols["small"]))
                    or num(cellv(rowvals, cols["large"]))
                    or num(cellv(rowvals, cols.get("case"))))
            weight = num(cellv(rowvals, cols["weight"]))
            # some tabs put GSM in the Weight column (e.g. Atelier Luxe "280")
            if not gsm and weight:
                try:
                    if float(weight) > 50:
                        gsm, weight = weight, ""
                except ValueError:
                    pass
            item_type, size = split_desc(desc)
            dims = fmt_dims(cellv(rowvals, cols["size"]))
            hem = cellv(rowvals, cols.get("hem"))
            # fitted items: append pocket depth to dimensions (39" x 75" x 9")
            # unless the size column already carried three dimensions
            if (hem and DEPTH_RE.match(hem) and "FITTED" in desc.upper()
                    and dims.count("x") < 2):
                dims = f"{dims} x {hem}"
            records.append({
                "sheet": ws.title,
                "block": block,
                "item_type": item_type,
                "size_type": size or cellv(rowvals, cols["size"]),
                "dimensions": dims,
                "cut_size": fmt_dims(cellv(rowvals, cols.get("cut"))),
                "hem_depth": hem,
                "gsm": re.sub(r"\s*GSM$", "", gsm, flags=re.I),
                "weight": weight,
                "weight_unit": cols.get("weight_unit", "") if weight else "",
                "case_pack": case,
                "color": cellv(rowvals, cols["color"]),
                "description": desc,
            })
    return records


def write_master(records, path="specs_master.csv"):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(records[0].keys()))
        w.writeheader()
        w.writerows(records)
    print(f"Wrote {path} ({len(records)} rows)")


def write_blocks(records, path="blocks.txt"):
    seen = OrderedDict()
    for r in records:
        if r["block"]:
            seen.setdefault(r["block"], r["sheet"])
    with open(path, "w", encoding="utf-8") as f:
        for block, sheet in seen.items():
            f.write(f"[tab {sheet}] {block}\n")
    print(f"Wrote {path} ({len(seen)} product blocks)")


def write_specs(records, mapping_path, path="specs.csv"):
    mapping = []
    with open(mapping_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            h, m = norm(row.get("handle")), norm(row.get("block_match"))
            if h and m:
                mapping.append((h, m.upper()))

    ALT_COLOR = re.compile(r"BONE|ECRU|NATURAL|IVORY|GREY|GRAY|KHAKI|TAN|LATTE", re.I)
    grouped = OrderedDict()  # key -> {row, colors:set}
    for r in records:
        blk = r["block"].upper()
        # longest (most specific) matching block_match wins, so a short
        # brand name like "Martex Simplicity" never steals rows from a more
        # specific "Martex Simplicity - Sheets"
        best = None
        for handle, match in mapping:
            if match.upper() in blk and (best is None or len(match) > len(best[1])):
                best = (handle, match)
        if best:
            handle, match = best
            if True:
                key = (handle, r["item_type"].upper(), r["size_type"].upper(),
                       r["gsm"], r["dimensions"].upper())
                g = grouped.setdefault(key, {"row": {
                    "handle": handle,
                    "item_type": r["item_type"],
                    "size": r["size_type"],
                    "gsm": r["gsm"],
                    "dimensions": r["dimensions"],
                    "weight": r["weight"],
                    "weight_unit": r.get("weight_unit", ""),
                    "case_pack": r["case_pack"],
                }, "colors": set()})
                if r["color"]:
                    g["colors"].add(r["color"])

    # star rows whose alt-color availability differs from the product norm:
    # if only SOME rows of a handle come in the alt color, star those rows;
    # if ALL rows do, no stars (the availability line covers it)
    by_handle = {}
    for key, g in grouped.items():
        by_handle.setdefault(key[0], []).append(g)
    out = []
    for handle, groups in by_handle.items():
        has_alt = [any(ALT_COLOR.search(c) for c in g["colors"]) for g in groups]
        star_some = any(has_alt) and not all(has_alt)
        for g, alt in zip(groups, has_alt):
            row = dict(g["row"])
            if star_some and alt:
                row["size"] = "*" + row["size"]
            out.append(row)

    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["handle", "item_type", "size", "gsm",
                                          "dimensions", "weight", "weight_unit",
                                          "case_pack"])
        w.writeheader()
        w.writerows(out)
    print(f"Wrote {path} ({len(out)} rows for {len(mapping)} mapped handles)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--map", default="mapping.csv")
    args = ap.parse_args()

    records = extract(args.workbook)
    if not records:
        print("No records extracted — check the workbook structure.")
        return 1
    write_master(records)
    write_blocks(records)

    import os
    if os.path.exists(args.map):
        write_specs(records, args.map)
    else:
        print(f"(no {args.map} found — create it from blocks.txt to generate specs.csv)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
